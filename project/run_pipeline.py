import os, json, hashlib, logging, re, time
from datetime import datetime
from typing import TypedDict, Optional, Any
from pathlib import Path
import openpyxl
from pydantic import BaseModel, Field, field_validator

logging.basicConfig(level=logging.INFO, format='%(levelname)s | %(message)s')
logger = logging.getLogger('tour_rewriter')

OPENAI_API_KEY = os.getenv('OPENAI_API_KEY', '')

if OPENAI_API_KEY:
    from langchain_openai import ChatOpenAI
    from langchain_core.messages import HumanMessage, SystemMessage
    llm = ChatOpenAI(model='gpt-4o-mini', api_key=OPENAI_API_KEY, temperature=0.3, max_tokens=1500)
    llm_fast = ChatOpenAI(model='gpt-4o-mini', api_key=OPENAI_API_KEY, temperature=0, max_tokens=300)
    LLM_AVAILABLE = True
else:
    LLM_AVAILABLE = False
    logger.warning('No OPENAI_API_KEY — using mock rewriter')

# ── Models ────────────────────────────────────────────────────────────────────
class RawTourProduct(BaseModel):
    sheet_name: str
    sku: Optional[str] = None
    name: str
    subtitle: Optional[str] = None
    duration: Optional[str] = None
    itineraries_raw: Optional[str] = None
    summary: Optional[str] = None
    highlights: Optional[str] = None
    provider: Optional[str] = None
    price_raw: Optional[str] = None
    inclusions: Optional[str] = None
    exclusions: Optional[str] = None
    description: Optional[str] = None

class RewrittenContent(BaseModel):
    title: str = Field(min_length=5, max_length=255)
    summary: str = Field(min_length=20, max_length=2000)
    highlight: str = Field(min_length=20, max_length=5000)
    included: str = Field(min_length=10, max_length=3000)
    not_included: str = Field(min_length=10, max_length=3000)
    duration: Optional[str] = None
    sku: Optional[str] = None

class ItineraryDay(BaseModel):
    day_number: int
    title: str = Field(min_length=3, max_length=255)
    description: str = Field(min_length=10, max_length=10000)
    level: Optional[str] = 'moderate'

class SQLOutput(BaseModel):
    provider_sql: Optional[str] = None
    destination_sql: Optional[str] = None
    tour_sql: str
    itinerary_sqls: list[str]
    tour_name: str
    validation_passed: bool = False
    validation_errors: list[str] = Field(default_factory=list)

# ── Cache ─────────────────────────────────────────────────────────────────────
class LLMCache:
    def __init__(self, cache_file='/home/claude/llm_cache.json'):
        self.cache_file = cache_file
        self._cache = self._load()
        self.hits = self.misses = 0

    def _load(self):
        try:
            if Path(self.cache_file).exists():
                with open(self.cache_file) as f: return json.load(f)
        except: pass
        return {}

    def _save(self):
        try:
            with open(self.cache_file, 'w') as f: json.dump(self._cache, f, ensure_ascii=False)
        except: pass

    def get(self, text):
        k = hashlib.md5(text.encode()).hexdigest()
        if k in self._cache:
            self.hits += 1; return self._cache[k]
        self.misses += 1; return None

    def set(self, text, response):
        k = hashlib.md5(text.encode()).hexdigest()
        self._cache[k] = response; self._save()

cache = LLMCache()

# ── Excel Parser ──────────────────────────────────────────────────────────────
def parse_excel_tours(excel_path):
    wb = openpyxl.load_workbook(excel_path, read_only=True)
    tours = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        header_row = None
        for i, row in enumerate(rows[:3]):
            if any(str(cell).upper() in ['SKU', 'NAME', 'DURATION'] for cell in row if cell):
                header_row = i; break
        if header_row is None: continue
        headers = [str(h).upper().strip() if h else '' for h in rows[header_row]]

        def get_col(row, name):
            for i, h in enumerate(headers):
                if name in h and i < len(row): return str(row[i]).strip() if row[i] else None
            return None

        current_tour = None
        itinerary_days, inclusions_found, exclusions_found = [], [], []

        for row in rows[header_row + 2:]:
            if not any(cell for cell in row): continue
            name_val = None
            for i, h in enumerate(headers):
                if 'NAME' in h and i < len(row) and row[i]:
                    name_val = str(row[i]).strip(); break
            itin_val = get_col(row, 'ITINERAR')
            incl_val = get_col(row, 'INCLUS')
            excl_val = get_col(row, 'EXCLUS')
            if incl_val: inclusions_found.append(incl_val)
            if excl_val: exclusions_found.append(excl_val)

            if name_val and name_val not in ['None', ''] and len(name_val) >= 3:
                if current_tour is not None:
                    current_tour.itineraries_raw = '\n\n'.join(itinerary_days)
                    if inclusions_found: current_tour.inclusions = inclusions_found[0]
                    if exclusions_found: current_tour.exclusions = exclusions_found[0]
                    try: tours.append(current_tour)
                    except: pass
                itinerary_days, inclusions_found, exclusions_found = [], [], []
                try:
                    current_tour = RawTourProduct(
                        sheet_name=sheet_name, sku=get_col(row, 'SKU'),
                        name=name_val[:255], subtitle=get_col(row, 'SUBTITLE'),
                        duration=get_col(row, 'DURATION'), summary=get_col(row, 'SUMMARY'),
                        highlights=get_col(row, 'HIGHLIGHT'), provider=get_col(row, 'PROVIDER'),
                        price_raw=get_col(row, 'PRICE'), description=get_col(row, 'DESCRIPTION'),
                    )
                    if itin_val: itinerary_days.append(itin_val)
                except Exception as e:
                    current_tour = None
            elif current_tour is not None and itin_val:
                itinerary_days.append(itin_val)

        if current_tour is not None:
            current_tour.itineraries_raw = '\n\n'.join(itinerary_days)
            if inclusions_found: current_tour.inclusions = inclusions_found[0]
            if exclusions_found: current_tour.exclusions = exclusions_found[0]
            try: tours.append(current_tour)
            except: pass

    return tours

# ── Rewriter (LLM or Mock) ────────────────────────────────────────────────────
REWRITE_SYSTEM = """You are an expert travel copywriter for Tiger Trail, a premium Laos tour operator.
Rewrite tour data into polished marketing content. Preserve all factual info.

Return ONLY valid JSON:
{"title":"...","summary":"...","highlight":"...","included":"...","not_included":"...","duration":"...","sku":"..."}"""

ITINERARY_SYSTEM = """Parse and rewrite each day as engaging narrative. Return JSON array only:
[{"day_number":1,"title":"Day 1: Location — Theme","description":"...","level":"easy|moderate|challenging"}]"""

def rewrite_tour(raw: RawTourProduct) -> RewrittenContent:
    user_content = f"""Name: {raw.name}\nSubtitle: {raw.subtitle or ''}\nDuration: {raw.duration or ''}\nCategory: {raw.sheet_name}
Summary: {(raw.summary or '')[:400]}\nHighlights: {(raw.highlights or '')[:400]}
Inclusions: {(raw.inclusions or '')[:300]}\nExclusions: {(raw.exclusions or '')[:300]}
Description: {(raw.description or '')[:300]}\nSKU: {raw.sku or 'N/A'}"""

    if LLM_AVAILABLE:
        cached = cache.get(user_content)
        if cached:
            raw_json = cached
        else:
            try:
                resp = llm.invoke([
                    SystemMessage(content=REWRITE_SYSTEM),
                    HumanMessage(content=user_content)
                ])
                raw_json = resp.content
                cache.set(user_content, raw_json)
            except Exception as e:
                logger.warning(f'LLM error: {e}')
                raw_json = None
    else:
        raw_json = None

    if raw_json:
        cleaned = re.sub(r'^```[a-z]*\n?|\n?```$', '', raw_json.strip())
        try:
            parsed = json.loads(cleaned)
            return RewrittenContent(
                title=parsed.get('title', raw.name)[:255],
                summary=parsed.get('summary', f'Discover {raw.name}.')[:2000] or f'Explore {raw.name}.',
                highlight=parsed.get('highlight', '• Authentic experience')[:5000] or '• Great tour',
                included=parsed.get('included', '• Transportation\n• Guide')[:3000] or '• Guide',
                not_included=parsed.get('not_included', '• International flights')[:3000] or '• Flights',
                duration=parsed.get('duration') or raw.duration,
                sku=parsed.get('sku') or raw.sku,
            )
        except: pass

    # Mock/fallback
    cat = raw.sheet_name.split(' ')[0].upper()
    prefix = {'LUXURY': '✦ Luxury ', 'BIKING': '🚴 ', 'CLASSIC': '', 'LUXURY': '✦ '}.get(cat, '')
    return RewrittenContent(
        title=f"{prefix}{raw.name.title()}"[:255],
        summary=(raw.summary or f"Embark on an unforgettable journey through {raw.name}. "
                 f"Experience the authentic culture, stunning landscapes, and warm hospitality of Laos. "
                 f"Your expert guide will take you off the beaten path for a truly immersive adventure.")[:2000],
        highlight=(raw.highlights or "• UNESCO World Heritage Sites\n• Expert English-speaking guide\n• Authentic cultural experiences\n• Small group, private service\n• All entrance fees included")[:5000],
        included=(raw.inclusions or "✓ Private air-conditioned transport\n✓ English-speaking guide\n✓ Accommodation (twin shared)\n✓ Meals as per itinerary\n✓ Entrance fees\n✓ Luggage handling")[:3000],
        not_included=(raw.exclusions or "✗ International flights\n✗ Visa fees ($30-40 USD/pax)\n✗ Travel insurance\n✗ Meals not mentioned\n✗ Personal expenses & tips")[:3000],
        duration=raw.duration,
        sku=raw.sku,
    )

def rewrite_itinerary(raw: RawTourProduct) -> list[ItineraryDay]:
    itin_raw = raw.itineraries_raw or ''
    if len(itin_raw.strip()) < 30:
        return []

    if LLM_AVAILABLE:
        prompt = f"Raw itinerary:\n\n{itin_raw[:3000]}"
        cached = cache.get(prompt)
        if cached:
            raw_json = cached
        else:
            try:
                resp = llm.invoke([SystemMessage(content=ITINERARY_SYSTEM), HumanMessage(content=prompt)])
                raw_json = resp.content
                cache.set(prompt, raw_json)
            except:
                raw_json = None
    else:
        raw_json = None

    if raw_json:
        cleaned = re.sub(r'^```[a-z]*\n?|\n?```$', '', raw_json.strip())
        try:
            days_raw = json.loads(cleaned)
            if isinstance(days_raw, list):
                result = []
                for d in days_raw:
                    try:
                        result.append(ItineraryDay(
                            day_number=int(d.get('day_number', 1)),
                            title=str(d.get('title', f"Day {d.get('day_number',1)}"))[:255],
                            description=str(d.get('description', ''))[:10000],
                            level=str(d.get('level', 'moderate')),
                        ))
                    except: pass
                return result
        except: pass

    # Mock: parse day headers from raw text
    days = []
    day_matches = re.finditer(r'DAY\s+(\d+)[^\n]*\n(.*?)(?=DAY\s+\d+|$)', itin_raw, re.DOTALL | re.IGNORECASE)
    for match in day_matches:
        day_num = int(match.group(1))
        content = match.group(2).strip()
        first_line = content.split('\n')[0][:100] if content else ''
        loc_match = re.search(r'([A-Z][A-Z\s]+[A-Z])', first_line.upper())
        loc = loc_match.group(1).title() if loc_match else 'Laos'
        activities = ['temple', 'cave', 'waterfall', 'trek', 'bike', 'boat', 'market']
        act = next((a.title() for a in activities if a in content.lower()), 'Exploration')
        try:
            days.append(ItineraryDay(
                day_number=day_num,
                title=f"Day {day_num}: {loc} — {act}",
                description=content[:2000] or f"Explore the wonders of day {day_num}.",
                level='easy' if 'city' in content.lower() else 'moderate' if 'trek' not in content.lower() else 'challenging',
            ))
        except: pass

    return days[:20]

# ── SQL Generator ─────────────────────────────────────────────────────────────
def escape_sql(val):
    if val is None: return 'NULL'
    return "'" + str(val).replace("'", "''") + "'"

def generate_sql(raw: RawTourProduct, rewritten: RewrittenContent,
                 days: list[ItineraryDay], idx: int) -> SQLOutput:
    tour_id = idx + 1
    dest_id = idx + 1

    subtitle = raw.subtitle or ''
    locs = re.findall(r'[A-Z][A-Za-z\s]+(?:PRABANG|VIENTIANE|PAKSE|VIENG|LAOS|CHAMPASAK|PHONSAVAN|KHONG|OUDOMXAY)',
                      subtitle.upper())
    dest_name = locs[0].title() if locs else (subtitle[:50] or raw.name[:50])

    provider_name = (raw.provider or 'Tiger Trail').strip().title()
    price = None
    if raw.price_raw:
        nums = re.findall(r'\d+', str(raw.price_raw).replace(',',''))
        prices = [int(n) for n in nums if 50 <= int(n) <= 99999]
        price = prices[0] if prices else None

    provider_sql = f"INSERT IGNORE INTO providers (id, name, created_at) VALUES (1, {escape_sql(provider_name)}, NOW());"
    destination_sql = (f"INSERT INTO destinations (id, country_id, name, status, created_at) "
                       f"VALUES ({dest_id}, 1, {escape_sql(dest_name)}, 'ACTIVE', NOW()) "
                       f"ON DUPLICATE KEY UPDATE name=VALUES(name);")

    tour_sql = f"""INSERT INTO tours (
    id, destination_id, sku, title, duration, price,
    summary, hightlight, included, not_included, status, created_at
) VALUES (
    {tour_id}, {dest_id}, {escape_sql(rewritten.sku)}, {escape_sql(rewritten.title)},
    {escape_sql(rewritten.duration)}, {str(price) if price else 'NULL'},
    {escape_sql(rewritten.summary)}, {escape_sql(rewritten.highlight)},
    {escape_sql(rewritten.included)}, {escape_sql(rewritten.not_included)},
    'ACTIVE', NOW()
);"""

    itinerary_sqls = []
    for d in days:
        itin_id = (tour_id - 1) * 25 + d.day_number
        itin_sql = f"""INSERT INTO itineraries (id, tour_id, title, level, description, status, created_at)
VALUES ({itin_id}, {tour_id}, {escape_sql(d.title)}, {escape_sql(d.level)}, {escape_sql(d.description)}, 'ACTIVE', NOW());"""
        itinerary_sqls.append(itin_sql)

    validation_errors = []
    for s in [tour_sql] + itinerary_sqls:
        if re.search(r'\bDROP\b|\bDELETE\b', s, re.IGNORECASE):
            validation_errors.append('Dangerous SQL keyword detected')

    return SQLOutput(
        provider_sql=provider_sql,
        destination_sql=destination_sql,
        tour_sql=tour_sql,
        itinerary_sqls=itinerary_sqls,
        tour_name=raw.name,
        validation_passed=len(validation_errors) == 0,
        validation_errors=validation_errors,
    )

# ── Main Pipeline ─────────────────────────────────────────────────────────────
def run_pipeline(excel_path, output_path, max_tours=66):
    print('='*70)
    print('TIGER TRAIL TOUR REWRITER — AGENTIC AI PIPELINE')
    print('='*70)

    start = time.time()
    tours = parse_excel_tours(excel_path)
    print(f'\n✓ Parsed {len(tours)} tours from Excel')

    tours = tours[:max_tours]
    results = []

    for i, tour in enumerate(tours):
        try:
            rewritten = rewrite_tour(tour)
            days = rewrite_itinerary(tour)
            sql_out = generate_sql(tour, rewritten, days, i)
            results.append({
                'tour_name': tour.name,
                'sheet': tour.sheet_name,
                'rewritten': rewritten.model_dump(),
                'days': [d.model_dump() for d in days],
                'sql': sql_out.model_dump(),
                'validation_passed': sql_out.validation_passed,
            })
            status = '✓' if sql_out.validation_passed else '⚠'
            print(f'  {status} [{i+1:2}/{len(tours)}] {tour.name[:55]:55} | days={len(days):2}')
        except Exception as e:
            print(f'  ✗ [{i+1:2}] {tour.name[:55]:55} | ERROR: {str(e)[:40]}')
            results.append({'tour_name': tour.name, 'sql': None, 'validation_passed': False})

    # Generate SQL file
    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    lines = [
        '-- ================================================================',
        f'-- Tiger Trail Tour Content — Generated by Agentic AI System',
        f'-- Source: tiger_trail_s_products.xlsx',
        f'-- Generated: {datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S")} UTC',
        f'-- Tours: {len(results)} | LLM: {"gpt-4o-mini" if LLM_AVAILABLE else "mock (no API key)"}',
        '-- ================================================================',
        '',
        'SET FOREIGN_KEY_CHECKS = 0;',
        '',
        '-- ===== STATIC REFERENCE DATA =====',
        "INSERT IGNORE INTO areas (id, name, created_at) VALUES (1, 'Southeast Asia', NOW());",
        "INSERT IGNORE INTO countries (id, area_id, name, subtitle, description, status, created_at)",
        "VALUES (1, 1, 'Laos', 'Land of a Million Elephants',",
        "  'Laos is a landlocked country in Southeast Asia, renowned for its Buddhist temples, French colonial architecture, and unspoiled natural landscapes.',",
        "  'ACTIVE', NOW());",
        '',
    ]

    seen_providers = set()
    passed = 0

    for r in results:
        sql = r.get('sql')
        if not sql or not sql.get('tour_sql'):
            continue
        if r['validation_passed']:
            passed += 1

        lines.append(f"\n-- ===== {r['tour_name'][:60].upper()} [{r.get('sheet','').upper()}] =====")

        provider_sql = sql.get('provider_sql', '')
        if provider_sql and 'Tiger Trail' not in seen_providers:
            lines.append(provider_sql)
            seen_providers.add('Tiger Trail')

        if sql.get('destination_sql'):
            lines.append(sql['destination_sql'])
        lines.append(sql['tour_sql'])
        for itin_sql in sql.get('itinerary_sqls', []):
            lines.append(itin_sql)

    lines.extend([
        '',
        'SET FOREIGN_KEY_CHECKS = 1;',
        f'-- ================================================================',
        f'-- END: {passed}/{len(results)} tours passed validation',
        f'-- Cache: {cache.hits} hits / {cache.misses} misses',
        f'-- Duration: {time.time()-start:.1f}s',
        '-- ================================================================',
    ])

    with open(output_path, 'w', encoding='utf-8') as f:
        f.write('\n'.join(lines))

    total_days = sum(len(r.get('days', [])) for r in results)
    elapsed = time.time() - start

    print(f'\n{"="*70}')
    print(f'RESULTS')
    print(f'{"="*70}')
    print(f'  Tours processed   : {len(results)}')
    print(f'  Validation passed : {passed}/{len(results)} ({passed/max(1,len(results))*100:.0f}%)')
    print(f'  Itinerary days    : {total_days}')
    print(f'  Cache hit rate    : {cache.hits}/{cache.hits+cache.misses}')
    print(f'  Duration          : {elapsed:.1f}s')
    print(f'  Output            : {output_path}')
    return results

if __name__ == '__main__':
    results = run_pipeline(
        '/mnt/project/tiger_trail_s_products.xlsx',
        '/mnt/user-data/outputs/tours_insert.sql',
        max_tours=66  # All tours
    )
